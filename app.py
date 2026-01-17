# =====================================================
# Sigma Consultants – Full CRM (FINAL STABLE BUILD)
# =====================================================

import streamlit as st
import pandas as pd
from datetime import datetime
import os
import io
from openpyxl.styles import Font, PatternFill

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="Sigma Consultants",
    layout="wide"
)

# ---------------- MOBILE TOGGLE ----------------
if "is_mobile" not in st.session_state:
    st.session_state.is_mobile = True

st.session_state.is_mobile = st.toggle(
    "📱 Mobile View",
    value=st.session_state.is_mobile
)
is_mobile = st.session_state.is_mobile

# ---------------- CARD UI ----------------
def card(title, value):
    st.markdown(
        f"""
        <div style="padding:14px;border-radius:12px;
        background:#ffffff;margin-bottom:10px;
        box-shadow:0 3px 8px rgba(0,0,0,0.08)">
        <div style="font-size:13px;color:#555">{title}</div>
        <div style="font-size:22px;font-weight:600">{value}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

# ---------------- LOGIN ----------------
PASSWORD = "sigma123"

if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    st.title("🔐 Sigma Consultants Login")
    pwd = st.text_input("Enter Password", type="password")

    if st.button("Login", use_container_width=True):
        if pwd == PASSWORD:
            st.session_state.auth = True
            st.session_state.page = "Summary"
            st.rerun()
        else:
            st.error("Incorrect password")

    st.stop()

# ---------------- FILES ----------------
CLIENT_FILE = "clients.xlsx"
PROPOSAL_FILE = "proposals.xlsx"

# ---------------- LOADERS ----------------
def load_clients():
    if os.path.exists(CLIENT_FILE):
        df = pd.read_excel(CLIENT_FILE)
    else:
        df = pd.DataFrame(columns=[
            "Client_ID","Client_Name",
            "Created_Date","Is_Archived","Notes"
        ])
        df.to_excel(CLIENT_FILE, index=False)

    if "Is_Archived" not in df.columns:
        df["Is_Archived"] = False
    if "Notes" not in df.columns:
        df["Notes"] = ""

    return df

def load_proposals():
    if os.path.exists(PROPOSAL_FILE):
        df = pd.read_excel(PROPOSAL_FILE)
    else:
        df = pd.DataFrame(columns=[
            "Proposal_ID","Client_ID","Client_Name",
            "Proposal_Cost","Rate","Final_Cost","Profit",
            "Start_Date","End_Date","Status","Closing_Date"
        ])
        df.to_excel(PROPOSAL_FILE, index=False)

    for c in ["Start_Date","End_Date","Closing_Date"]:
        df[c] = pd.to_datetime(df[c], errors="coerce")

    return df

# ---------------- SESSION DATA ----------------
if "clients_df" not in st.session_state:
    st.session_state.clients_df = load_clients()

if "proposals_df" not in st.session_state:
    st.session_state.proposals_df = load_proposals()

clients_df = st.session_state.clients_df
proposals_df = st.session_state.proposals_df

def save_clients():
    clients_df.to_excel(CLIENT_FILE, index=False)

def save_proposals():
    proposals_df.to_excel(PROPOSAL_FILE, index=False)

def new_client_id():
    return f"SIG-C-{int(datetime.now().timestamp())}"

def new_proposal_id():
    return f"SIG-P-{int(datetime.now().timestamp())}"

# ---------------- PAGE STATE ----------------
if "page" not in st.session_state:
    st.session_state.page = "Summary"

# ---------------- SIDEBAR ----------------
with st.sidebar:
    st.markdown("## 📂 Sigma Consultants")

    menu = [
        ("📊 Summary","Summary"),
        ("📥 Export Data","Export"),
    ]

    for label, page in menu:
        if st.button(label, use_container_width=True):
            st.session_state.page = page
            st.rerun()

    st.markdown("---")
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.auth = False
        st.session_state.page = "Summary"
        st.rerun()

# =====================================================
# ================= SUMMARY PAGE ======================
# =====================================================
if st.session_state.page == "Summary":

    st.header("📊 Summary")

    if proposals_df.empty:
        st.info("No proposals available")
        st.stop()

    total_inv = proposals_df["Proposal_Cost"].sum()
    total_profit = proposals_df["Profit"].sum()
    open_cnt = len(proposals_df[proposals_df["Status"] == "Open"])

    if is_mobile:
        card("Investment", f"₹ {total_inv:,.2f}")
        card("Profit", f"₹ {total_profit:,.2f}")
        card("Open Proposals", open_cnt)
    else:
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Investment", f"₹ {total_inv:,.2f}")
        c2.metric("Total Profit", f"₹ {total_profit:,.2f}")
        c3.metric("Open Proposals", open_cnt)

    st.markdown("---")
    st.subheader("📅 Proposal Details")
    st.caption("ℹ️ Amounts shown in actual ₹ (rupees)")

    status_options = ["All"] + sorted(
        proposals_df["Status"].dropna().unique().tolist()
    )
    selected_status = st.selectbox("Select Status", status_options)

    df = proposals_df.copy()
    if selected_status != "All":
        df = df[df["Status"] == selected_status]

    if df.empty:
        st.warning("No data available")
        st.stop()

    end_dates = sorted(df["End_Date"].dropna().unique())
    selected_end_date = st.selectbox(
        "Select End Date",
        end_dates,
        format_func=lambda x: x.strftime("%d-%m-%Y")
    )

    df = df[df["End_Date"] == selected_end_date]
    df["Rate_Int"] = df["Rate"].round(0).astype(int)

    g_inv = df["Proposal_Cost"].sum()
    g_final = df["Final_Cost"].sum()
    g_profit = df["Profit"].sum()

    st.markdown("### 🔢 End Date Totals")

    if is_mobile:
        card("Investment", f"₹ {g_inv:,.2f}")
        card("Final Amount", f"₹ {g_final:,.2f}")
        card("Profit", f"₹ {g_profit:,.2f}")
    else:
        g1, g2, g3 = st.columns(3)
        g1.metric("Investment", f"₹ {g_inv:,.2f}")
        g2.metric("Final Amount", f"₹ {g_final:,.2f}")
        g3.metric("Profit", f"₹ {g_profit:,.2f}")

    st.markdown("---")

    summary_df = (
        df.groupby(["Rate_Int","Start_Date"], as_index=False)
        .agg({
            "Proposal_Cost":"sum",
            "Final_Cost":"sum",
            "Profit":"sum"
        })
        .sort_values(["Rate_Int","Start_Date"])
    )

    for _, row in summary_df.iterrows():
        st.markdown(
            f"""
**📅 {row['Start_Date'].strftime('%d-%m-%Y')} | {row['Rate_Int']} %**  
💰 Investment : ₹ {row['Proposal_Cost']:,.2f}  
📈 Final : ₹ {row['Final_Cost']:,.2f}  
📊 Profit : ₹ {row['Profit']:,.2f}
"""
        )
        st.markdown("---")

# =====================================================
# ================= EXPORT PAGE =======================
# =====================================================
if st.session_state.page == "Export":

    st.header("📤 Export Data")

    if proposals_df.empty:
        st.info("No data available")
        st.stop()

    status_options = ["All"] + sorted(
        proposals_df["Status"].dropna().unique().tolist()
    )
    export_status = st.selectbox("Select Status", status_options)

    df = proposals_df.copy()
    if export_status != "All":
        df = df[df["Status"] == export_status]

    min_date = df["Start_Date"].min().date()
    max_date = df["End_Date"].max().date()

    date_range = st.date_input(
        "Select Date Range",
        value=(min_date, max_date)
    )

    start_date, end_date = map(pd.to_datetime, date_range)
    df = df[
        (df["Start_Date"] >= start_date) &
        (df["End_Date"] <= end_date)
    ]

    client_options = ["All"] + sorted(df["Client_Name"].dropna().unique())
    client = st.selectbox("Select Client", client_options)

    if client != "All":
        df = df[df["Client_Name"] == client]

    df["Rate"] = df["Rate"].round(0).astype(int)

    export_df = df[
        ["Client_Name","Start_Date",
         "Proposal_Cost","Rate",
         "Final_Cost","Profit"]
    ].copy()

    export_df["Start_Date"] = export_df["Start_Date"].dt.strftime("%d-%m-%Y")

    total_row = {
        "Client_Name":"GRAND TOTAL",
        "Start_Date":"",
        "Proposal_Cost":export_df["Proposal_Cost"].sum(),
        "Rate":"",
        "Final_Cost":export_df["Final_Cost"].sum(),
        "Profit":export_df["Profit"].sum()
    }

    export_df = pd.concat(
        [export_df, pd.DataFrame([total_row])],
        ignore_index=True
    )

    def export_excel(df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            ws = writer.sheets["Sheet1"]

            last_row = ws.max_row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFF4CCCC",
                    end_color="FFF4CCCC",
                    fill_type="solid"
                )
        output.seek(0)
        return output

    st.download_button(
        "⬇️ Download Excel",
        data=export_excel(export_df),
        file_name="Sigma_Export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
